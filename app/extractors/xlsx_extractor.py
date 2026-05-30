"""XLSX (Excel 2007+) document extraction using openpyxl.

Extracts text content from all sheets, rows, and cells.
Handles merged cells, formula results (not formulas), dates as strings.
"""

from io import BytesIO
from typing import Optional

from ..extractor import ExtractionResult
from .base import BaseExtractor


class XlsxExtractor(BaseExtractor):
    """Extractor for .xlsx (Office Open XML Spreadsheet) documents.

    Uses openpyxl to iterate all sheets, rows, and cells, extracting
    text content formatted as:
        [Sheet: SheetName]
        Cell A1: value
        Cell A2: value
    """

    def extract(self, data: bytes, filename: Optional[str] = None) -> ExtractionResult:
        """Extract text from an .xlsx spreadsheet.

        Args:
            data: Raw .xlsx file bytes.
            filename: Optional original filename.

        Returns:
            ExtractionResult with extracted text content.
        """
        try:
            import openpyxl  # type: ignore[import-not-found]
        except ImportError:
            return ExtractionResult(
                full_text="",
                confidence=0.0,
                requirements=["openpyxl package is not installed"],
            )

        try:
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
        except Exception:
            return ExtractionResult(
                full_text="",
                confidence=0.0,
                requirements=["Failed to open XLSX workbook"],
            )

        all_text_parts: list[str] = []
        total_cells = 0
        populated_cells = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_parts: list[str] = []
            sheet_cells = 0
            sheet_populated = 0

            for row in ws.iter_rows():
                row_parts: list[str] = []
                for cell in row:
                    total_cells += 1
                    sheet_cells += 1
                    if cell.value is not None:
                        sheet_populated += 1
                        populated_cells += 1
                        # Convert to string — openpyxl returns native Python types
                        cell_str = str(cell.value).strip()
                        if cell_str:
                            row_parts.append(f"Cell {cell.coordinate}: {cell_str}")

                if row_parts:
                    sheet_parts.extend(row_parts)

            if sheet_parts:
                all_text_parts.append(f"[Sheet: {sheet_name}]")
                all_text_parts.extend(sheet_parts)

        wb.close()

        if not all_text_parts:
            return ExtractionResult(
                full_text="",
                confidence=0.0,
                requirements=["No data found in XLSX workbook"],
            )

        full_text = "\n".join(all_text_parts)

        # Confidence based on content volume
        if populated_cells == 0:
            confidence = 0.0
        elif total_cells == 0:
            confidence = 0.0
        else:
            ratio = populated_cells / max(total_cells, 1)
            if ratio >= 0.5 and populated_cells >= 20:
                confidence = 0.7
            elif populated_cells >= 10:
                confidence = 0.5
            elif populated_cells >= 3:
                confidence = 0.3
            else:
                confidence = 0.1

        # Description: first meaningful content
        description = ""
        for line in full_text.splitlines():
            if line.startswith("Cell ") and len(line) > 60:
                description = line
                break

        return ExtractionResult(
            full_text=full_text,
            description=description,
            confidence=confidence,
            raw_text_preview=full_text[:500] if full_text else None,
        )
